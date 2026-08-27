from __future__ import annotations

import csv
from dataclasses import dataclass
from math import cos, radians, sin, sqrt
from pathlib import Path
import sqlite3
from typing import Any

import numpy as np


@dataclass(frozen=True)
class ImuSample:
    timestamp_s: float
    quaternion_wxyz: tuple[float, float, float, float]
    acceleration_xyz: tuple[float, float, float] | None = None
    gyro_xyz: tuple[float, float, float] | None = None


@dataclass(frozen=True)
class SlamImuData:
    samples: list[ImuSample]
    frame_times_s: list[float]
    frame_sensor_timestamps_ns: list[int]
    frame_codec_pts_us: list[int]
    timestamp_origin_ns: int
    metadata: dict[str, str]
    rolling_shutter_skew_s: float | None
    gyro_filter_window_s: float


def _col(row: dict[str, Any], *names: str) -> Any | None:
    lowered = {k.lower().strip(): v for k, v in row.items()}
    for name in names:
        if name in lowered and lowered[name] not in ("", None):
            return lowered[name]
    return None


def _load_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("Reading XLSX IMU files requires openpyxl.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        names = [str(value).strip() if value is not None else "" for value in header]
        return [
            {name: value for name, value in zip(names, row)}
            for row in rows
            if row is not None and any(value is not None for value in row)
        ]

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader)


def _normalize(q: np.ndarray) -> np.ndarray:
    norm = np.linalg.norm(q)
    if norm == 0:
        return np.array([1.0, 0.0, 0.0, 0.0])
    return q / norm


def _quat_mul(a: np.ndarray, b: np.ndarray) -> np.ndarray:
    aw, ax, ay, az = a
    bw, bx, by, bz = b
    return np.array(
        [
            aw * bw - ax * bx - ay * by - az * bz,
            aw * bx + ax * bw + ay * bz - az * by,
            aw * by - ax * bz + ay * bw + az * bx,
            aw * bz + ax * by - ay * bx + az * bw,
        ],
        dtype=float,
    )


def _delta_from_gyro(gx: float, gy: float, gz: float, dt: float, units: str) -> np.ndarray:
    if units == "deg_s":
        gx, gy, gz = radians(gx), radians(gy), radians(gz)
    omega = np.array([gx, gy, gz], dtype=float)
    angle = float(np.linalg.norm(omega) * dt)
    if angle == 0:
        return np.array([1.0, 0.0, 0.0, 0.0])
    axis = omega / np.linalg.norm(omega)
    half = angle / 2.0
    return _normalize(np.array([cos(half), *(axis * sin(half))], dtype=float))


def _axis_rotation_matrix(axis_rotation: list[float] | tuple[float, ...] | None) -> np.ndarray:
    if not axis_rotation:
        return np.eye(3, dtype=float)
    matrix = np.array(axis_rotation, dtype=float)
    if matrix.size != 9:
        raise ValueError("IMU axis rotation must contain 9 numbers.")
    return matrix.reshape(3, 3)


def _rotate_vector(vector: tuple[float, float, float], matrix: np.ndarray) -> tuple[float, float, float]:
    rotated = matrix @ np.array(vector, dtype=float)
    return (float(rotated[0]), float(rotated[1]), float(rotated[2]))


def load_imu_csv(
    path: str | Path,
    gyro_units: str = "rad_s",
    axis_rotation: list[float] | tuple[float, ...] | None = None,
    gyro_scale: float = 1.0,
) -> list[ImuSample]:
    rows = _load_rows(Path(path))

    if not rows:
        raise ValueError("IMU file is empty.")

    samples: list[ImuSample] = []
    q = np.array([1.0, 0.0, 0.0, 0.0], dtype=float)
    last_t: float | None = None
    imu_to_camera = _axis_rotation_matrix(axis_rotation)

    for row in rows:
        t_raw = _col(row, "timestamp_s", "time_s", "t", "time", "timestamp")
        t_us_raw = _col(row, "timestamp(us)", "timestamp_us", "time_us", "timestamp_usec")
        if t_raw is None:
            if t_us_raw is None:
                raise ValueError("IMU CSV needs a timestamp_s/time_s/t or timestamp(us) column.")
            t = float(t_us_raw) / 1_000_000.0
        else:
            t = float(t_raw)

        ax = _col(row, "acc_x", "accel_x", "accelerometer_x", "ax")
        ay = _col(row, "acc_y", "accel_y", "accelerometer_y", "ay")
        az = _col(row, "acc_z", "accel_z", "accelerometer_z", "az")
        acceleration = None
        if ax is not None and ay is not None and az is not None:
            acceleration = _rotate_vector((float(ax), float(ay), float(az)), imu_to_camera)

        qw = _col(row, "qw", "quat_w", "w")
        qx = _col(row, "qx", "quat_x", "x")
        qy = _col(row, "qy", "quat_y", "y")
        qz = _col(row, "qz", "quat_z", "z")
        if all(v is not None for v in (qw, qx, qy, qz)):
            q = _normalize(np.array([float(qw), float(qx), float(qy), float(qz)], dtype=float))
        else:
            gx = _col(row, "gx", "gyro_x", "omega_x")
            gy = _col(row, "gy", "gyro_y", "omega_y")
            gz = _col(row, "gz", "gyro_z", "omega_z")
            if gx is None or gy is None or gz is None:
                raise ValueError("IMU CSV needs either qw/qx/qy/qz or gx/gy/gz columns.")
            gyro = _rotate_vector((float(gx), float(gy), float(gz)), imu_to_camera)
            gyro = tuple(value * float(gyro_scale) for value in gyro)
            dt = 0.0 if last_t is None else max(0.0, t - last_t)
            q = _normalize(_quat_mul(q, _delta_from_gyro(*gyro, dt, gyro_units)))
        if all(v is not None for v in (qw, qx, qy, qz)):
            gyro = None

        samples.append(
            ImuSample(
                timestamp_s=t,
                quaternion_wxyz=tuple(float(v) for v in q),
                acceleration_xyz=acceleration,
                gyro_xyz=gyro,
            )
        )
        last_t = t

    return samples


def _interpolate_timed_vector(
    rows: list[tuple[int, tuple[float, float, float]]],
    timestamp_ns: int,
    cursor: int,
) -> tuple[tuple[float, float, float] | None, int]:
    if not rows:
        return None, cursor
    while cursor + 1 < len(rows) and rows[cursor + 1][0] <= timestamp_ns:
        cursor += 1
    if cursor + 1 >= len(rows):
        return rows[-1][1], cursor
    before_t, before = rows[cursor]
    after_t, after = rows[cursor + 1]
    if timestamp_ns <= before_t or after_t <= before_t:
        return before, cursor
    alpha = max(0.0, min(1.0, (timestamp_ns - before_t) / (after_t - before_t)))
    return tuple(before[i] + (after[i] - before[i]) * alpha for i in range(3)), cursor


def _centered_vector_average(
    samples: list[tuple[int, tuple[float, float, float]]],
    window_s: float,
) -> list[tuple[int, tuple[float, float, float]]]:
    if len(samples) < 2 or window_s <= 0.0:
        return samples
    half_window_ns = int(window_s * 1_000_000_000.0 / 2.0)
    prefix = [[0.0, 0.0, 0.0]]
    for _, vector in samples:
        previous = prefix[-1]
        prefix.append([previous[i] + vector[i] for i in range(3)])
    averaged: list[tuple[int, tuple[float, float, float]]] = []
    left = 0
    right = 0
    for timestamp_ns, _ in samples:
        while left < len(samples) and samples[left][0] < timestamp_ns - half_window_ns:
            left += 1
        while right < len(samples) and samples[right][0] <= timestamp_ns + half_window_ns:
            right += 1
        count = max(1, right - left)
        averaged.append(
            (
                timestamp_ns,
                tuple((prefix[right][axis] - prefix[left][axis]) / count for axis in range(3)),
            )
        )
    return averaged


def load_slamimu(
    path: str | Path,
    axis_rotation: list[float] | tuple[float, ...] | None = None,
    gyro_filter_window_s: float = 0.015,
) -> SlamImuData:
    """Load the SLAM Camera SQLite sidecar using Camera2 time as the origin."""

    database = Path(path)
    connection = sqlite3.connect(f"{database.resolve().as_uri()}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    try:
        integrity = connection.execute("PRAGMA integrity_check").fetchone()
        if not integrity or integrity[0] != "ok":
            raise ValueError(f"SLAM IMU database integrity check failed: {integrity[0] if integrity else 'unknown'}")

        metadata = {
            str(row["key"]): str(row["value"])
            for row in connection.execute("SELECT key, value FROM session_metadata")
        }
        frames = connection.execute(
            """
            SELECT frame_number, sensor_timestamp_ns, codec_pts_us, rolling_shutter_skew_ns
            FROM video_frames
            ORDER BY frame_number
            """
        ).fetchall()
        if not frames:
            raise ValueError("SLAM IMU database has no video frame timing records.")
        frame_numbers = [int(row["frame_number"]) for row in frames]
        if any(b != a + 1 for a, b in zip(frame_numbers, frame_numbers[1:])):
            raise ValueError("SLAM IMU video frame numbers are not contiguous.")

        origin_ns = int(frames[0]["sensor_timestamp_ns"])
        frame_timestamps_ns = [int(row["sensor_timestamp_ns"]) for row in frames]
        frame_times_s = [(timestamp - origin_ns) / 1_000_000_000.0 for timestamp in frame_timestamps_ns]
        frame_pts_us = [int(row["codec_pts_us"]) for row in frames]
        if any(b <= a for a, b in zip(frame_timestamps_ns, frame_timestamps_ns[1:])):
            raise ValueError("SLAM IMU video timestamps are not strictly increasing.")

        accel_rows: list[tuple[int, tuple[float, float, float]]] = []
        for row in connection.execute(
            """
            SELECT timestamp_ns, x, y, z, bias_x, bias_y, bias_z
            FROM imu_samples
            WHERE sensor_type = 'ACCEL_UNCALIBRATED'
            ORDER BY timestamp_ns
            """
        ):
            accel_rows.append(
                (
                    int(row["timestamp_ns"]),
                    (
                        float(row["x"]) - float(row["bias_x"] or 0.0),
                        float(row["y"]) - float(row["bias_y"] or 0.0),
                        float(row["z"]) - float(row["bias_z"] or 0.0),
                    ),
                )
            )

        gyro_rows = connection.execute(
            """
            SELECT timestamp_ns, x, y, z, bias_x, bias_y, bias_z
            FROM imu_samples
            WHERE sensor_type = 'GYRO_UNCALIBRATED'
            ORDER BY timestamp_ns
            """
        ).fetchall()
        if len(gyro_rows) < 2:
            raise ValueError("SLAM IMU database needs at least two uncalibrated gyro samples.")

        imu_to_camera = _axis_rotation_matrix(axis_rotation)
        gyro_vectors = [
            (
                int(row["timestamp_ns"]),
                _rotate_vector(
                    (
                        float(row["x"]) - float(row["bias_x"] or 0.0),
                        float(row["y"]) - float(row["bias_y"] or 0.0),
                        float(row["z"]) - float(row["bias_z"] or 0.0),
                    ),
                    imu_to_camera,
                ),
            )
            for row in gyro_rows
        ]
        gyro_vectors = _centered_vector_average(gyro_vectors, gyro_filter_window_s)
        samples: list[ImuSample] = []
        q = np.array([1.0, 0.0, 0.0, 0.0], dtype=float)
        last_timestamp_ns: int | None = None
        accel_cursor = 0
        for timestamp_ns, gyro in gyro_vectors:
            if last_timestamp_ns is not None and timestamp_ns <= last_timestamp_ns:
                raise ValueError("SLAM IMU gyro timestamps are not strictly increasing.")
            acceleration_sensor, accel_cursor = _interpolate_timed_vector(
                accel_rows,
                timestamp_ns,
                accel_cursor,
            )
            acceleration = (
                _rotate_vector(acceleration_sensor, imu_to_camera)
                if acceleration_sensor is not None
                else None
            )
            dt = (
                0.0
                if last_timestamp_ns is None
                else (timestamp_ns - last_timestamp_ns) / 1_000_000_000.0
            )
            q = _normalize(_quat_mul(q, _delta_from_gyro(*gyro, dt, "rad_s")))
            samples.append(
                ImuSample(
                    timestamp_s=(timestamp_ns - origin_ns) / 1_000_000_000.0,
                    quaternion_wxyz=tuple(float(value) for value in q),
                    acceleration_xyz=acceleration,
                    gyro_xyz=gyro,
                )
            )
            last_timestamp_ns = timestamp_ns

        skew_values = [
            int(row["rolling_shutter_skew_ns"])
            for row in frames
            if row["rolling_shutter_skew_ns"] is not None
        ]
        rolling_shutter_skew_s = (
            sum(skew_values) / len(skew_values) / 1_000_000_000.0
            if skew_values
            else None
        )
        return SlamImuData(
            samples=samples,
            frame_times_s=frame_times_s,
            frame_sensor_timestamps_ns=frame_timestamps_ns,
            frame_codec_pts_us=frame_pts_us,
            timestamp_origin_ns=origin_ns,
            metadata=metadata,
            rolling_shutter_skew_s=rolling_shutter_skew_s,
            gyro_filter_window_s=gyro_filter_window_s,
        )
    finally:
        connection.close()


def summarize_rate(samples: list[ImuSample]) -> float:
    if len(samples) < 2:
        return 0.0
    deltas = [b.timestamp_s - a.timestamp_s for a, b in zip(samples, samples[1:]) if b.timestamp_s > a.timestamp_s]
    if not deltas:
        return 0.0
    return 1.0 / (sum(deltas) / len(deltas))
